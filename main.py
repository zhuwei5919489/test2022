
import re
import openpyxl as pxl

"""读取题目"""

f = open("C:\\Users\\zhuwei\\Desktop\\智慧学习平台考试\\test.txt","r", encoding="utf-8")  #打开d盘test文件夹内名为test的文本文件
txt=""

for line in f.readlines():
    txt += line

new_txt = txt.replace(" ","").replace("\n","")

lst,lst0,t1,t2,t3 = [],[],[],[],[]
st = ()
timu=""
sn = 0
# print(new_txt)


# text0 = "[saadmin].([\u4E00-\u9FA5]+).*?</span>"

text0 = "]([\u4E00-\u9FA5]+)\d"

# text1 = "<divstyle='height:5px;margin-top:-5px'></div><b>\d.(.*?)</b>"     ##匹配题目
text1 = "><b>\d.(.*?)</b>"     ##匹配题目

# text1 = "></span></span><br><b>\d.(.*?)</b>"     ##匹配题目(退出后重进  循环记得len(t0)-1)


text3 = "var  tm  = '([a-zA-Z0-9]+)'"  ### 匹配题目编号

text4 = "</b><br>(.*?)</div>"     ### 匹配题目选项


t0 = re.findall(text0,new_txt)   #正则表达式，匹配题目类型
t1 = re.findall(text1,new_txt)   #正则表达式，匹配题目
t2 = re.findall(text3,txt)       #匹配题目编号
t3 = re.findall(text4,txt)       #匹配选项


# print(t0)

lst_option = []
option = ""
options= []
for i in t3:
    j = i.split("<br>")
    # print(j)
    for k in j:
        option +=k+"&&"
    # print(option)
    lst_option.append(option)
    option = ""

# print(lst_option)
#
# print(len(lst_option))




lst_sn = []

for num in t2[0].split('s'):
    if sn != "":
        lst_sn.append(num)

# print(lst_sn)


# for obj in t0:
#     print(obj)

if t1 != [] :
    text2 = "&#\d+;([a-zA-Z0-9_\u4e00-\u9fa5\u3002\uff1b\uff0c\uff1a\u201c\u201d\uff08\uff09\u3001\uff1f\u300a\u300b\(\)].*?)&#"

    ###   \u4e00-\u9fa5 匹配中文
    ###   \u3002\uff1b\uff0c\uff1a\u201c\u201d\uff08\uff09\u3001\uff1f\u300a\u300b  匹配标点符号。 ；  ， ： “ ”（ ） 、 ？ 《 》




    for i in range(0,len(t1)):
        # print(t1[i])
        t2 =  re.findall(text2,t1[i],re.DOTALL)
        # print(t2)
        s= []
        for k in t2:   # 数组去重
            if k  not in s:
                s.append(k)
        sn +=1
        for j in s:
            timu += j

        print(str(sn)+"."+timu)

        lst0.append(timu)
        lst.append(t0[i]+timu)

        timu=""
    st = ()

f.close()           # 将文本关闭
# print(lst)
# print(lst[47])

print("合计提取到题目 %d 个。" % sn)

"""匹配题库"""

book = pxl.load_workbook("C:\\Users\\zhuwei\\Desktop\\智慧学习平台考试\\智慧学习平台考试答案匹配2022.xlsx",data_only=True)    #打开test1.xlsx
sheet1 = book["答案"]        #选取工作表test1
sheet1['A1'].value = "序号"   # 对单元格A1赋值
sheet1['B1'].value = "答案-基于题目"   # 对单元格B1赋值
sheet1['C1'].value = "答案-基于题目&选项"
sheet1['D1'].value = "答案-基于编号"
sheet1['E1'].value = "题目类型"
sheet1['F1'].value = "题目"   # 对单元格C1赋值
sheet1['G1'].value = "选项A"   #
sheet1['H1'].value = "选项B"   #
sheet1['I1'].value = "选项C"   #
sheet1['J1'].value = "选项D"   #
sheet1['K1'].value = "选项E"   #
sheet1['L1'].value = "选项F"   #
sheet1['M1'].value = "选项G"   #
sheet1['N1'].value = "选项H"   #




for row in sheet1['A2:O200']:    #将指定区域'A2:C200'清空
  for cell in row:
    cell.value = None

for i in range(1,len(lst)+1):     #将题目写入Excel
    sheet1.cell(i+1,1).value = i
    sheet1.cell(i+1,5).value = t0[i-1]
    sheet1.cell(i+1,6).value = lst0[i-1]

for i in range(1,len(lst_option)+1):  #将选项写入EXCEL

    # print(lst_option[i - 1])

    options = lst_option[i-1].split("&&")
    # print(lst_option[i-1])
    # print(options)
    # print(len(options))

    for j in range(1,len(options)+1):
        sheet1.cell(i + 1, 6+j).value = options[j - 1]
    #
    # options = []

lst_string = []     ### 题目+题目类型+选项字符串

string = ""

for i in range(1,len(lst)+1):     #将题目写入Excel

    for j in range(1,12):
        if sheet1.cell(i+1,4+j).value != None:
             string += str(sheet1.cell(i+1,4+j).value)
    lst_string.append(string)

    string = ""







sheet2 = book["财务题库"]                                            #选取的考试题库工作表名称
# sheet2 = book["1"]

sheet3= book["Sheet1"]

# str11 = sheet3.cell(2,4).value
# str12 = sheet3.cell(2,9).value
#
# print(str11)
# print(str11.replace(" ",""))
# print(str12)

lst1,dt = [],{}

lst2,dt_sn = [],{}

lst3,dt_answer = [],{}

answer_opt = ""

n = sheet2.max_row           #tiku的最大行数

# print(n)
for i in range(2,n+2):       #将题目及答案写入列表lst1

    tiku = str(sheet2.cell(i,3).value).replace(" ","").replace("\n","")+str(sheet2.cell(i,4).value).replace(" ","").replace("\n","")
    answer_1 = (tiku,sheet2.cell(i,2).value)
    lst1.append(answer_1)

    answer_2 = (str(sheet2.cell(i,1).value),sheet2.cell(i,2).value)
    lst2.append(answer_2)


    for j in range(1,10):
        if sheet2.cell(i,4+j).value != None:
            answer_opt += sheet2.cell(i,4+j).value

    # print(answer_opt)

    answer_3 = (str(tiku+answer_opt), sheet2.cell(i, 2).value)
    lst3.append(answer_3)
    answer_opt = ""


# print(len(lst1))
#
# print(len(lst3))


dt = dict(lst1)              #将列表lst1转换为字典dt

dt_sn = dict(lst2)

dt_answer = dict(lst3)
# print(dt_sn)
# print(lst_sn)

# print(lst1)


flag1 = 0
flag2 = 0
flag3 = 0




# print(dt)
for i in range(1,len(lst)+1):        #将答案返回并写入第二列

    if  lst[i-1] in dt:            ##  基于题目匹配
        sheet1.cell(i+1,2).value = dt[lst[i-1]]
        flag1 += 1
    else:
        sheet1.cell(i + 1, 2).value = ""



    if  lst_sn[i-1] in dt_sn:     ##  基于编号匹配

        sheet1.cell(i+1,4).value = dt_sn[lst_sn[i-1]]
        flag2 += 1
    else:
        sheet1.cell(i + 1, 4).value = ""



    if lst_string[i-1] in dt_answer:     ## 基于题目及选项匹配
        sheet1.cell(i+1,3).value = dt_answer[lst_string[i-1]]
        flag3 += 1
    else:
        sheet1.cell(i + 1, 3).value = ""






print("\n")
print("题目合计 %d 个,基于(题目)匹配到答案的题目 %d 个,未匹配到答案题目 %d 个! ^-^ " % (sn,flag1,sn-flag1))
print("\n")
print("题目合计 %d 个,基于(题目+选项)匹配到答案的题目 %d 个,未匹配到答案题目 %d 个! ^-^ " % (sn,flag3,sn-flag3))

print("\n")
print("题目合计 %d 个,基于(编号)匹配到答案的题目 %d 个,未匹配到答案题目 %d 个! ^-^ " % (sn,flag2,sn-flag2))

book.save("C:\\Users\\zhuwei\\Desktop\\智慧学习平台考试\\智慧学习平台考试答案匹配2022.xlsx")

























